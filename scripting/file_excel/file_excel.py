import openpyxl

file_path = "data/octobre.xlsx"
wb = openpyxl.load_workbook(file_path)
# print(wb.sheetnames)
# sheet = wb[wb.sheetnames[0]]
sheet = wb.active
# cell = sheet["A1"]
# print(sheet.max_column)
# print(sheet.max_row)
for row in range(2, sheet.max_row):
    cell = sheet.cell(row, 2)
    if not cell.value is None:
        print(cell.value)
