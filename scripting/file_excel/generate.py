import openpyxl
from openpyxl.chart import BarChart, Series, Reference

wb1 = openpyxl.load_workbook("data/octobre.xlsx", data_only=True)
wb2 = openpyxl.load_workbook("data/novembre.xlsx", data_only=True)
wb3 = openpyxl.load_workbook("data/decembre.xlsx", data_only=True)

def add_data_wb(wb, data):
    sheet = wb.active
    for row in range(2, sheet.max_row):
        article_name = sheet.cell(row, 1).value
        if not article_name:
            break
        price = sheet.cell(row, 4).value
        if data.get(article_name):
            data[article_name].append(price)
        else:
            data[article_name] = [price]

data = {}
add_data_wb(wb1, data)
add_data_wb(wb2, data)
add_data_wb(wb3, data)

# print(data)

wb_output = openpyxl.Workbook()
sheet = wb_output.active
sheet["A1"] = "Article"
sheet["B1"] = "Octobre"
sheet["C1"] = "November"
sheet["D1"] = "December"

row = 2
for i in data.items():
    # print(i)
    article_name = i[0]
    sales = i[1]
    sheet.cell(row, 1).value = article_name
    for j in range(0, len(sales)):
        sheet.cell(row, j+2).value = sales[j]
    row += 1

chart_reference = Reference(sheet, min_col=2, min_row=2, max_col=sheet.max_column, max_row=2)
chart_series = Series(chart_reference, title="Total ventes €")
chart = BarChart()
chart.title = "Evolution du prix des pommes"
chart.append(chart_series)
sheet.add_chart(chart, "F2")

wb_output.save("data/output.xlsx")

